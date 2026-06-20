# file: app/routers/admin_router.py

import asyncio
import json
from datetime import datetime, timezone
from io import BytesIO
from typing import List

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, status, BackgroundTasks
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from starlette.responses import StreamingResponse

from app.database import get_db, AsyncSessionLocal  # get_db باید نسخه async باشد
from app import crud, schemas, models
from app.state import live_states
from app.websocket_manager import manager

router = APIRouter(
    prefix="/api/admin",
    tags=["Admin"],
    responses={404: {"description": "Not found"}},
)




async def send_questions_timed(public_id: str):
    """تابع پس‌زمینه برای ارسال زمان‌بندی شده سوالات (نسخه اصلاح شده)."""
    async with AsyncSessionLocal() as db:
        try:
            livestream = await crud.get_livestream_for_task(db, public_id=public_id)

            if not livestream:
                print(f"Livestream {public_id} not found for background task.")
                return

            questions = sorted(livestream.questions, key=lambda q: q.display_time_seconds)
            start_time = datetime.now(timezone.utc)

            for question in questions:
                await db.refresh(livestream, attribute_names=['status'])

                if livestream.status != models.LiveStreamStatus.ACTIVE:
                    print(f"Livestream {public_id} stopped. Ending task.")
                    break

                sleep_duration = question.display_time_seconds - (
                        datetime.now(timezone.utc) - start_time).total_seconds()
                if sleep_duration > 0:
                    await asyncio.sleep(sleep_duration)

                if livestream.status != models.LiveStreamStatus.ACTIVE:
                    break

                options_for_client = [{"id": opt.id, "text": opt.text} for opt in question.options]

                message = {
                    "type": "show_question",
                    "data": {
                        "id": question.id,
                        "text": question.text,
                        "question_type": question.question_type.value,
                        "options": options_for_client,
                        "duration": question.duration
                    }
                }

                # 👑 قدم اول و طلایی: ذخیره سوال فعلی در حافظه موقت سرور
                # با این کار، هر کسی در طول زمان نمایش این سوال ریکانکت شود، سرور می‌داند چه چیزی به او نشان دهد.
                live_states[public_id] = message

                # حالا با خیال راحت پیام را برای آنلاین‌ها بفرستید
                await manager.broadcast(json.dumps(message), public_id)

                # منتظر می‌مانیم تا زمان پاسخ‌گویی به سوال تمام شود
                await asyncio.sleep(question.duration)

                # 👑 قدم دوم: چون زمان سوال تمام شده، وضعیت را از حافظه پاک یا ریسیت می‌کنیم
                if public_id in live_states:
                    del live_states[public_id] # یا می‌توانید یک وضعیت خالی بگذارید

                hide_message = {"type": "hide_question", "data": {"question_id": question.id}}
                await manager.broadcast(json.dumps(hide_message), public_id)

        except Exception as e:
            import traceback
            print(f"Error in background task for {public_id}: {e}")
            traceback.print_exc()


@router.post("/livestreams", response_model=schemas.LiveStreamAdmin, status_code=status.HTTP_201_CREATED)
async def create_new_livestream(livestream: schemas.LiveStreamCreate, db: AsyncSession = Depends(get_db)):
    """ایجاد لایو. تسک پس‌زمینه در این مرحله اجرا نمی‌شود."""
    return await crud.create_livestream(db=db, livestream=livestream)


@router.get("/livestreams/{admin_secret}", response_model=schemas.LiveStreamAdmin)
async def read_livestream_for_admin(admin_secret: str, db: AsyncSession = Depends(get_db)):
    db_livestream = await crud.get_livestream_by_admin_secret(db, admin_secret=admin_secret)
    if db_livestream is None:
        raise HTTPException(status_code=404, detail="Livestream not found")
    return db_livestream


@router.get("/livestreams", response_model=List[schemas.LiveStreamForShow])
async def read_livestreams(db: AsyncSession = Depends(get_db)):
    db_livestreams = await crud.get_livestreams(db)

    # اگر می‌خواهید اگر لیست خالی بود خطا بدهید (اختیاری):
    if not db_livestreams:
        print("No livestreams found.")
        return []  # یا می‌توانید به جای خطا دادن، لیست خالی برگردانید که رایج‌تر است
    print(db_livestreams[0].title)
    return db_livestreams

@router.put("/livestreams/{admin_secret}", response_model=schemas.LiveStreamAdmin)
async def update_existing_livestream(admin_secret: str, livestream_update: schemas.LiveStreamCreate,
                                     db: AsyncSession = Depends(get_db)):
    """ویرایش لایو. تسک پس‌زمینه در این مرحله اجرا نمی‌شود."""
    db_livestream = await crud.get_livestream_by_admin_secret(db, admin_secret=admin_secret)
    if db_livestream is None:
        raise HTTPException(status_code=404, detail="Livestream not found")
    return await crud.update_livestream(db=db, db_livestream=db_livestream, livestream_update=livestream_update)

@router.delete("/livestreams/delete/{admin_secret}", response_model=bool)
async def delete_livestream(admin_secret: str,
                                     db: AsyncSession = Depends(get_db)):
    """ویرایش لایو. تسک پس‌زمینه در این مرحله اجرا نمی‌شود."""
    db_livestream = await crud.delete_livestream(db, admin_secret=admin_secret)
    if not db_livestream:
        raise HTTPException(status_code=404, detail="Livestream not found")
    return True

@router.put("/livestreams/active/{admin_secret}", response_model=bool)
async def active_livestream(admin_secret: str,
                                     db: AsyncSession = Depends(get_db)):
    """ویرایش لایو. تسک پس‌زمینه در این مرحله اجرا نمی‌شود."""
    db_livestream = await crud.active_livestream(db, admin_secret=admin_secret)
    if not db_livestream:
        raise HTTPException(status_code=404, detail="Livestream not found")
    return True

@router.put("/livestreams/deactive/{admin_secret}", response_model=bool)
async def active_livestream(admin_secret: str,
                                     db: AsyncSession = Depends(get_db)):
    """ویرایش لایو. تسک پس‌زمینه در این مرحله اجرا نمی‌شود."""
    db_livestream = await crud.de_active_livestream(db, admin_secret=admin_secret)
    if not db_livestream:
        raise HTTPException(status_code=404, detail="Livestream not found")
    return True

@router.post("/livestreams/{admin_secret}/start", status_code=status.HTTP_200_OK)
async def start_livestream(admin_secret: str, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_db)):
    """شروع لایو و اجرای تسک پس‌زمینه برای ارسال سوالات."""
    livestream = await crud.get_livestream_by_admin_secret(db, admin_secret=admin_secret)
    if not livestream:
        raise HTTPException(status_code=404, detail="Livestream not found")
    if livestream.status == models.LiveStreamStatus.ACTIVE:
        raise HTTPException(status_code=400, detail="Livestream is already active.")

    livestream.status = models.LiveStreamStatus.ACTIVE
    await db.commit()

    background_tasks.add_task(send_questions_timed, livestream.public_id)
    return {"message": "Livestream started."}


@router.post("/livestreams/{admin_secret}/stop", status_code=status.HTTP_200_OK)
async def stop_livestream(admin_secret: str, db: AsyncSession = Depends(get_db)):
    """متوقف کردن لایو."""
    livestream = await crud.get_livestream_by_admin_secret(db, admin_secret=admin_secret)
    if not livestream:
        raise HTTPException(status_code=404, detail="Livestream not found")
    if livestream.status == models.LiveStreamStatus.FINISHED:
        raise HTTPException(status_code=400, detail="Livestream is already finished.")

    livestream.status = models.LiveStreamStatus.FINISHED
    await db.commit()
    live_states.pop(livestream.public_id, None)
    end_message = {"type": "end_live", "data": {"message": "این لایو به پایان رسید."}}
    await manager.broadcast(json.dumps(end_message), livestream.public_id)
    return {"message": "Livestream stopped."}

@router.delete("/livestreams/delete-answers/{admin_secret}", response_model=bool)
async def delete_livestream_answers(admin_secret: str,
                                     db: AsyncSession = Depends(get_db)):
    """ویرایش لایو. تسک پس‌زمینه در این مرحله اجرا نمی‌شود."""
    db_livestream = await crud.delete_livestream_answers_and_users(db, admin_secret=admin_secret)
    if not db_livestream:
        raise HTTPException(status_code=404, detail="Livestream not found")
    return True

@router.post("/livestreams/{admin_secret}/pend", status_code=status.HTTP_200_OK)
async def pend_livestream(admin_secret: str, db: AsyncSession = Depends(get_db)):
    livestream = await crud.get_livestream_by_admin_secret(db, admin_secret=admin_secret)
    if not livestream:
        raise HTTPException(status_code=404, detail="Livestream not found")
    if livestream.status == models.LiveStreamStatus.PENDING:
        raise HTTPException(status_code=400, detail="Livestream is already pended.")

    livestream.status = models.LiveStreamStatus.PENDING
    await db.commit()
    return {"message": "Livestream pended."}


@router.get("/export-results/{livestream_id}")
async def export_results(livestream_id: int, db: AsyncSession = Depends(get_db)):
    # ۱. دریافت سوالات لایو استریم
    q_stmt = select(models.Question).where(models.Question.livestream_id == livestream_id).order_by(
        models.Question.id.asc())
    questions_result = await db.execute(q_stmt)
    questions = questions_result.scalars().all()

    # ۲. دریافت کاربران به همراه پاسخ‌ها
    u_stmt = select(models.User).where(models.User.livestream_id == livestream_id).options(
        selectinload(models.User.answers).selectinload(models.Answer.selected_option)
    )
    users_result = await db.execute(u_stmt)
    users = users_result.scalars().all()

    # ۳. دریافت سوالات به همراه گزینه‌ها (برای شماره‌گذاری گزینه‌ها)
    q_with_options_stmt = select(models.Question).where(models.Question.livestream_id == livestream_id).options(
        selectinload(models.Question.options)
    ).order_by(models.Question.id.asc())

    questions_with_options_result = await db.execute(q_with_options_stmt)
    questions_with_options = questions_with_options_result.scalars().all()

    # مرتب‌سازی گزینه‌ها در پایتون
    for q in questions_with_options:
        q.options.sort(key=lambda x: x.id)

    questions_map = {q.id: q for q in questions_with_options}

    if not users:
        raise HTTPException(status_code=404, detail="کاربری برای این لایو استریم یافت نشد.")

    # ۴. پردازش داده‌ها
    data_for_export = []

    for index, user in enumerate(users, start=1):
        answers_map = {ans.question_id: ans for ans in user.answers}

        row = {
            "ردیف": index,
            "نام و نام خانوادگی": user.full_name,
            "شماره تماس": getattr(user, 'phone_number', '-'),  # اضافه کردن شماره تماس
        }

        user_score = 0

        for i,q in enumerate(questions,start=1):
            ans = answers_map.get(q.id)
            column_name = f"سوال {i}"

            if not ans:
                row[column_name] = "-"
            else:
                if q.question_type == models.QuestionType.MULTIPLE_CHOICE:
                    if ans.selected_option:
                        selected_option_id = ans.selected_option.id
                        question_obj = questions_map.get(q.id)
                        option_number = ""
                        if question_obj and question_obj.options:
                            for j, opt in enumerate(question_obj.options):
                                if opt.id == selected_option_id:
                                    option_number = str(j + 1)
                                    break
                        row[column_name] = option_number if option_number else ans.selected_option.text

                        if ans.selected_option.is_correct:
                            user_score += 1
                    else:
                        row[column_name] = "بدون پاسخ"
                else:
                    row[column_name] = ans.answer_text if ans.answer_text else "بدون پاسخ"

        row["امتیاز"] = user_score
        data_for_export.append(row)

    # ۵. تولید فایل اکسل
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'نتایج آزمون'

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    rtl_alignment = Alignment(horizontal='right', vertical='center', readingOrder=2)

    if data_for_export:
        headers = list(data_for_export[0].keys())
        sheet.append(headers)

        # استایل سرصفحه
        for col_idx, header_cell in enumerate(sheet[1], 1):
            header_cell.font = Font(bold=True)
            header_cell.border = thin_border
            header_cell.alignment = rtl_alignment
            sheet.column_dimensions[get_column_letter(col_idx)].width = 20  # تنظیم عرض پیش‌فرض

        # نوشتن داده‌ها
        for row_idx, row_dict in enumerate(data_for_export, start=2):
            row_values = [str(row_dict.get(header, "")) for header in headers]
            sheet.append(row_values)

            for col_idx, cell in enumerate(sheet[row_idx], 1):
                cell.border = thin_border
                cell.alignment = rtl_alignment

    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=results_{livestream_id}.xlsx"}
    )
